import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles.borders import Border, Side
from openpyxl.cell import MergedCell
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import pdfplumber
import re
from xlwings import Book
import os
from datetime import datetime

# Define the folder where the PDF files are located
folder_path = "C:/Users/Twincom PC/Desktop/pdf_excel_upload"

# Function to automatically find the PDF file with the desired format
def find_pdf_file(folder_path):
    # List all files in the folder
    files = os.listdir(folder_path)
    # Filter for files matching the desired format
    pdf_files = [f for f in files if f.endswith(".pdf") and re.match(r".+\s\(.+\s+\d{1,2},\s+\d{4}\)\.pdf", f)]
    
    if not pdf_files:
        raise FileNotFoundError("No PDF files matching the expected format were found in the folder.")
    
    # If multiple files match, return the first (or handle as needed)
    return pdf_files[0]

# Debugging: Automatically detect the PDF file
try:
    pdf_file = find_pdf_file(folder_path)
    print(f"Detected PDF file: {pdf_file}")
except FileNotFoundError as e:
    print(e)
    pdf_file = None

# Proceed if a valid PDF file is detected
with pdfplumber.open(pdf_file) as f:
    all_tables = []
    for page in f.pages:
        tables = page.extract_tables()
        for table in tables:
            all_tables.append(table)

# Convert the tables into a DataFrame
df_list = [pd.DataFrame(table) for table in all_tables]
if len(df_list) > 1:
    df = pd.concat(df_list, ignore_index=True)
else:
    df = df_list[0]

# Clean and format the DataFrame
df.columns = df.iloc[0]
df = df[1:]
df.columns = df.columns.str.strip()
df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
df['Expense Type'] = df['Expense Type'].str.replace('\n', '', regex=False)
df['Expense Type'] = df['Expense Type'].str.replace(r'Reimbursement-|Liquidation-|Replenishment-', '', regex=True)
df['Expense Type'] = df['Expense Type'].str.upper()
df['Establishment'] = df['Establishment'].str.replace('\n', '', regex=False)
df['Establishment'] = df['Establishment'].str.upper()
df['PURPOSE'] = df['Expense Type'] + '\n' + df['Notes']

# Group by 'Date' and calculate the sum of 'Amount'
df_grouped = df.groupby('Date', as_index=False).agg({'Amount': 'sum'})
df_merged = df.merge(df_grouped, on='Date', how='left', suffixes=('', '_TOTAL'))

# Rename columns
df_merged.rename(columns={
    'Date': 'DATE',
    'Establishment': 'ESTABLISHMENT NAME',
    'OR No. / Ref.\nNo.': 'REF NO',
    'Amount': 'AMOUNT',
    'Amount_TOTAL': 'TOTAL',
    'Client': 'PROJECT NAME'
}, inplace=True)
df_merged['DATE'] = pd.to_datetime(df_merged['DATE']).dt.strftime('%m/%d/%Y')
df_merged['PCV No.'] = ''
df_merged['PROJECT CODE'] = ''
df_merged['PROJECT NAME'] = df_merged.get('PROJECT NAME', '')
df_merged['PO NUMBER'] = ''

# Drop unnecessary columns and rearrange columns
df_merged.drop(columns=['Expense Type', 'Time', 'Reimbursable', 'Notes'], inplace=True)
df_merged = df_merged[['PCV No.', 'DATE', 'ESTABLISHMENT NAME', 'REF NO', 'AMOUNT', 'TOTAL', 'PROJECT CODE', 'PROJECT NAME', 'PO NUMBER', 'PURPOSE']]

# Save the DataFrame to an Excel file
output_excel = "4_final_output.xlsx"
wb_new = Workbook()
ws_new = wb_new.active
ws_new.append(df_merged.columns.tolist())
for r in dataframe_to_rows(df_merged, index=False, header=False):
    ws_new.append(r)
wb_new.save(output_excel)

# Open the target Excel file and write the data into it
excel_file = 'Reimbursement_Final_File_2.xlsx'
wb = Book(excel_file)
ws = wb.sheets['EXPENSE FORM']

# Load the new data from the Excel file created
df_new_data = pd.read_excel(output_excel)

START_COL = 17
ADD_COL = df_new_data.shape[0] - 1

END_COL = START_COL + ADD_COL - 1

# Insert space for new data
ws.range((START_COL, 1), (START_COL + ADD_COL - 1, ws.cells.last_cell.column)).api.Insert()

# Write new data to the Excel sheet
ws.range(f"A{START_COL - 1}").value = df_new_data.values.tolist()

# Delete the unnecessary total row
ws.range((START_COL + ADD_COL - 1, 1), (END_COL, ws.cells.last_cell.column)).api.Delete()

# Initially save the excel file to be able to shift in openpyxl for font modifications
wb.save('Reimbursement_Init.xlsx')
app = wb.app
wb.close()
app.kill()

tempo_excel_file = 'Reimbursement_Init.xlsx'
wb_edit = load_workbook(tempo_excel_file)
ws_edit = wb_edit.active

def get_init_date(ws, END_COL):
    start_date = ws["B16"].value
    end_date = ws[f"B{END_COL - 1}"].value

    # Format the day parts to two digits
    start_day = start_date.strftime('%d')
    end_day = end_date.strftime('%d')
    end_year = end_date.strftime('%Y')
    end_month = end_date.strftime('%m')

    return start_day, end_day, end_month, end_year

# Function to merge cells in the Date and TOTAL columns
def merge_date_and_total(ws, date_col, total_col, start_row, date_data, total_data):
    current_date = date_data.iloc[0]
    current_total = total_data.iloc[0]
    start_merge = start_row

    for row in range(1, len(date_data)):
        if date_data.iloc[row] != current_date:  # When the date changes
            # Merge Date cells
            ws.merge_cells(start_row=start_merge, start_column=date_col, end_row=row + start_row - 1, end_column=date_col)
            ws.cell(row=start_merge, column=date_col).alignment = Alignment(horizontal="center", vertical="center")
            
            # Merge TOTAL cells corresponding to the merged Date cells
            ws.merge_cells(start_row=start_merge, start_column=total_col, end_row=row + start_row - 1, end_column=total_col)
            ws.cell(row=start_merge, column=total_col).value = current_total  # Set the total value
            ws.cell(row=start_merge, column=total_col).alignment = Alignment(horizontal="center", vertical="center")
            
            current_date = date_data.iloc[row]
            current_total = total_data.iloc[row]
            start_merge = row + start_row

    # Merge the last group of Date and TOTAL cells
    ws.merge_cells(start_row=start_merge, start_column=date_col, end_row=len(date_data) + start_row - 1, end_column=date_col)
    ws.cell(row=start_merge, column=date_col).alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells(start_row=start_merge, start_column=total_col, end_row=len(date_data) + start_row - 1, end_column=total_col)
    ws.cell(row=start_merge, column=total_col).alignment = Alignment(horizontal="center", vertical="center")


# Function to format cells with Arial 14
def format_cells(ws, font_name='Arial', font_size=14, bold_column=None):
    # Apply Arial 16 to all cells
    for row in ws.iter_rows():
        for cell in row:
            # Skip if the cell is a MergedCell object
            if not isinstance(cell, MergedCell):
                continue

            cell.font = Font(name=font_name, size=font_size)
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

             # Apply currency format to AMOUNT and TOTAL columns
            if cell.column == 5 or cell.column == 6:  # Column 5 ('E') for AMOUNT and column 6 ('F') for TOTAL
                cell.number_format = '"PHP" #,##0.00'  # PHP symbol with two decimal places

# Apply bold formatting to the specified column (for PURPOSE)
    if bold_column:
        for row in ws.iter_rows(min_col=bold_column, max_col=bold_column):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    continue
                cell.font = Font(name=font_name, size=font_size, bold=True)   


# Function to set column widths based on content
def auto_adjust_column_width(ws, min_row, max_row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col).column_letter  # Get the column letter
        
        for row in range(min_row, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                # Calculate the maximum length of the value in the column
                max_length = max(max_length, len(str(cell.value)))

        # Set the width with some padding
        adjusted_width = max_length + 2  # Add padding for better readability
        ws.column_dimensions[column_letter].width = adjusted_width


def adjust_cell_height(ws, min_row, max_row, min_col, max_col, default_char_width = 7):

    for row_h in range(min_row, max_row + 1):
        max_lines = 1  # Reset max_lines for each row

        for col_h in range(min_col, max_col + 1):
            cell_h = ws.cell(row=row_h, column=col_h)
            if cell_h.value:
                text = str(cell_h.value)
                # Get column width, defaulting to `default_char_width` if not set
                column_letter = get_column_letter(col_h)
                col_width = ws.column_dimensions[column_letter].width or default_char_width

                # Calculate maximum line length for wrapping
                max_line_length = int(col_width / default_char_width)
                if max_line_length == 0:  # Avoid division by zero
                    max_line_length = 1

                # Count explicit lines and wrapped lines
                wrapped_lines = sum((len(line) // max_line_length) + 1 for line in text.splitlines())
                max_lines = max(max_lines, wrapped_lines)

        # Adjust row height based on the total number of lines
        ws.row_dimensions[row_h].height = max_lines * 6

# Function fo fill borders
def create_border(ws, start_col, end_col):
    cell_range_body = ws[f'B{start_col}':f'I{end_col}']
    cell_range_left = ws[f'A{start_col}':f'A{end_col}']
    cell_range_right = ws[f'J{start_col}':f'J{end_col}']

    border_body = Border(left = Side(style = 'thin'),
                    right = Side(style = 'thin'),
                    top = Side(style = 'thin'),
                    bottom = Side(style = 'thin'))
    
    border_outer_left = Border(left = Side(style = 'medium'),
                    right = Side(style = 'thin'),
                    top = Side(style = 'thin'),
                    bottom = Side(style = 'thin'))

    border_outer_right = Border(left = Side(style = 'thin'),
                    right = Side(style = 'medium'),
                    top = Side(style = 'thin'),
                    bottom = Side(style = 'thin'))
    
    for row_body in cell_range_body:
        for cell_body in row_body:
            cell_body.border = border_body
    
    for row_left in cell_range_left:
        for cell_left in row_left:
            cell_left.border = border_outer_left

    for row_right in cell_range_right:
        for cell_right in row_right:
            cell_right.border = border_outer_right

def merge_name_cell(ws, start_row, start_column, end_row, end_column, pdf_file):
    match = re.search(r"(\w+)-(\w+ \w+)\s+\((\w+)\s+(\d{1,2}),\s+(\d{4})\)", pdf_file)
    if match:
        last_name = match.group(1).upper()
        first_name = match.group(2).upper()
        month_str = match.group(3)
        year = match.group(5)

        # Get today's day number
        today_day = datetime.today().strftime("%d")

    ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
    ws.cell(row=start_row, column=start_column).value = f"{first_name} {last_name}\n474"
    ws.cell(row=start_row, column=start_column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws['H7'] = f"{today_day}-{month_str[:3]}-{year[-2:]}"
    ws[f"E{END_COL+1}"] = f"=SUM(E{START_COL-1}:E{END_COL-1})"
    ws[f"F{END_COL+1}"] = f"=E{END_COL+1}"


start_day, end_day, end_month, end_year = get_init_date(ws_edit, END_COL)

# Apply the merging function on both the Date and TOTAL columns
merge_date_and_total(ws_edit, date_col=2, total_col=6, start_row=START_COL-1, date_data=df_merged['DATE'], total_data=df_merged['TOTAL'])

# Apply the formatting to the worksheet
format_cells(ws_edit) 

merge_name_cell(ws_edit, start_row=16, start_column=1, end_row=END_COL-1, end_column=1, pdf_file=pdf_file)

# Apply auto-adjust column widths
auto_adjust_column_width(ws_edit, min_row=1, max_row=10, min_col=16, max_col=END_COL-1)

# Apply auto-adjust row height
adjust_cell_height(ws_edit, min_row=16, max_row=END_COL-1, min_col=1, max_col=10)

# Apply border
create_border(ws_edit, start_col=16, end_col=END_COL-1)

# Extract details from the `pdf_file` name
def generate_output_filename(pdf_file, start_day, end_day, end_month, end_year):
    match = re.search(r"(\w+)-(\w+ \w+)\s+\((\w+)\s+(\d{1,2}),\s+(\d{4})\)", pdf_file)
    if match:
        last_name = match.group(1)
        first_name = match.group(2)
        month_str = match.group(3)

        # Combine into desired format
        output_filename = f"Reimbursement_{first_name} {last_name}_{end_month}.{start_day}-{end_day}.{end_year[-2:]}.xlsx"
    
        return output_filename
    else:
        raise ValueError("Filename format does not match expected pattern.")

# Generate output filename
output_excel_filename = generate_output_filename(pdf_file, start_day, end_day, end_month, end_year)

# Save and close the Excel workbook
wb_edit.save(output_excel_filename)
wb_edit.close()

# Function to delete files if they exist
def delete_file(file_path):
    try:
        if os.path.exists(file_path):
            os.remove(file_path)
            print(f"Deleted file: {file_path}")
        else:
            print(f"File not found: {file_path}")
    except Exception as e:
        print(f"Error deleting file {file_path}: {e}")

# Example usage
temp_files = ["4_final_output.xlsx", "Reimbursement_Init.xlsx"]

for temp_file in temp_files:
    delete_file(temp_file)
