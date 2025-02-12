import os
import pdfplumber
import pandas as pd
import re
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.cell import MergedCell
from openpyxl.utils import get_column_letter
import xlwings as xw
import pythoncom

EXCEL_TEMPLATE = "Reimbursement_Final_File_2.xlsx"

def process_pdf(pdf_file, output_folder):
    pythoncom.CoInitialize()
    with pdfplumber.open(pdf_file) as f:
        all_tables = []
        for page in f.pages:
            tables = page.extract_tables()
            for table in tables:
                all_tables.append(table)

    df_list = [pd.DataFrame(table) for table in all_tables]
    df = pd.concat(df_list, ignore_index=True) if len(df_list) > 1 else df_list[0]

    # Clean and transform DataFrame
    df.columns = df.iloc[0]
    df = df[1:]
    df.columns = df.columns.str.strip()
    df['Amount'] = pd.to_numeric(df['Amount'], errors='coerce')
    df['Expense Type'] = df['Expense Type'].str.replace('\n', '', regex=False)
    df['Expense Type'] = df['Expense Type'].str.replace(r'\n|Reimbursement-|Liquidation-|Replenishment-', '', regex=True).str.upper()
    df['Establishment'] = df['Establishment'].str.replace('\n', '', regex=False).str.upper()
    df['PURPOSE'] = df['Expense Type'] + '\n' + df['Notes']

    # Group by 'Date' and calculate sum of 'Amount'
    df_grouped = df.groupby('Date', as_index=False).agg({'Amount': 'sum'})
    df_merged = df.merge(df_grouped, on='Date', how='left', suffixes=('', '_TOTAL'))

    # Rename columns
    df_merged.rename(columns={
        'Date': 'DATE',
        'Establishment': 'ESTABLISHMENT NAME',
        'OR No. / Ref.\nNo.': 'REF NO',
        'Amount': 'AMOUNT',
        'Amount_TOTAL': 'TOTAL',
        'Client': 'PROJECT NAME'
    }, inplace=True)

    df_merged['DATE'] = pd.to_datetime(df_merged['DATE']).dt.strftime('%m/%d/%Y')
    df_merged['PCV No.'] = ''
    df_merged['PROJECT CODE'] = ''
    df_merged['PROJECT NAME'] = df_merged.get('PROJECT NAME', '')
    df_merged['PO NUMBER'] = ''

    # Drop unnecessary columns and rearrange columns
    df_merged.drop(columns=['Expense Type', 'Time', 'Reimbursable', 'Notes'], inplace=True, errors='ignore')
    df_merged = df_merged[['PCV No.', 'DATE', 'ESTABLISHMENT NAME', 'REF NO', 'AMOUNT', 'TOTAL', 'PROJECT CODE', 'PROJECT NAME', 'PO NUMBER', 'PURPOSE']]

    # Save the DataFrame to an Excel file
    output_excel = "4_final_output.xlsx"
    wb_new = Workbook()
    ws_new = wb_new.active
    ws_new.append(df_merged.columns.tolist())
    for r in dataframe_to_rows(df_merged, index=False, header=False):
        ws_new.append(r)
    wb_new.save(output_excel)

    # Open the target Excel file and write the data into it
    app = xw.App(visible=False)
    wb = app.books.open(EXCEL_TEMPLATE)
    ws = wb.sheets['EXPENSE FORM']

    # Load the new data from the Excel file created
    df_new_data = pd.read_excel(output_excel)

    START_COL = 17
    ADD_COL = df_new_data.shape[0] - 1
    END_COL = START_COL + ADD_COL - 1

    ws.range((START_COL, 1), (START_COL + ADD_COL - 1, ws.cells.last_cell.column)).api.Insert()
    ws.range(f"A{START_COL - 1}").value = df_new_data.values.tolist()
    ws.range((START_COL + ADD_COL - 1, 1), (END_COL, ws.cells.last_cell.column)).api.Delete()

    temp_excel_file = os.path.join(output_folder, "Reimbursement_Init.xlsx")
    wb.save(temp_excel_file)
    wb.close()
    app.kill()

    return finalize_excel(temp_excel_file, pdf_file, output_folder, df_merged, START_COL, END_COL)


def finalize_excel(excel_file, pdf_file, output_folder, df_merged, START_COL, END_COL):
    pythoncom.CoInitialize()
    wb = load_workbook(excel_file)
    ws = wb.active

    start_day, end_day, start_month, end_month, end_year = get_init_date(ws, END_COL)
    
    merge_date_and_total(ws, date_col=2, total_col=6, start_row=START_COL-1, date_data=df_merged['DATE'], total_data=df_merged['TOTAL'])
    format_cells(ws)
    merge_name_cell(ws, start_row=16, start_column=1, end_row=END_COL-1, end_column=1, pdf_file=pdf_file, START_COL=START_COL, END_COL=END_COL)
    auto_adjust_column_width(ws, min_row=1, max_row=10, min_col=16, max_col=END_COL-1)
    adjust_cell_height(ws, min_row=16, max_row=END_COL-1, min_col=1, max_col=10)
    create_border(ws, start_col=16, end_col=END_COL-1)

    output_filename = generate_output_filename(pdf_file, start_day, end_day, start_month, end_month, end_year)
    output_path = os.path.join(output_folder, output_filename)

    wb.save(output_path)
    wb.close()
    os.remove(excel_file)  # Cleanup temp file

    return output_path


# Supporting Functions
def get_init_date(ws, END_COL):
    start_date = ws["B16"].value
    end_date = ws[f"B{END_COL - 1}"].value
    return start_date.strftime('%d'), end_date.strftime('%d'), start_date.strftime('%m'), end_date.strftime('%m'), end_date.strftime('%Y')

def merge_date_and_total(ws, date_col, total_col, start_row, date_data, total_data):
    current_date = date_data.iloc[0]
    current_total = total_data.iloc[0]
    start_merge = start_row

    for row in range(1, len(date_data)):
        if date_data.iloc[row] != current_date:
            ws.merge_cells(start_row=start_merge, start_column=date_col, end_row=row + start_row - 1, end_column=date_col)
            ws.cell(row=start_merge, column=date_col).alignment = Alignment(horizontal="center", vertical="center")
            ws.merge_cells(start_row=start_merge, start_column=total_col, end_row=row + start_row - 1, end_column=total_col)
            ws.cell(row=start_merge, column=total_col).value = current_total # Set the total value
            ws.cell(row=start_merge, column=total_col).alignment = Alignment(horizontal="center", vertical="center")
            current_date, current_total, start_merge = date_data.iloc[row], total_data.iloc[row], row + start_row
    
    # Merge the last group of Date and TOTAL cells
    ws.merge_cells(start_row=start_merge, start_column=date_col, end_row=len(date_data) + start_row - 1, end_column=date_col)
    ws.cell(row=start_merge, column=date_col).alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells(start_row=start_merge, start_column=total_col, end_row=len(date_data) + start_row - 1, end_column=total_col)
    ws.cell(row=start_merge, column=total_col).alignment = Alignment(horizontal="center", vertical="center")

# Function to format cells with Arial 14
def format_cells(ws, font_name='Arial', font_size=14, bold_column=None):
    # Apply Arial 16 to all cells
    for row in ws.iter_rows():
        for cell in row:
            # Skip if the cell is a MergedCell object
            if not isinstance(cell, MergedCell):
                continue

            cell.font = Font(name=font_name, size=font_size)
            cell.alignment = Alignment(wrap_text=True, horizontal='left', vertical='center')

             # Apply currency format to AMOUNT and TOTAL columns
            if cell.column == 5 or cell.column == 6:  # Column 5 ('E') for AMOUNT and column 6 ('F') for TOTAL
                cell.number_format = '"PHP" #,##0.00'  # PHP symbol with two decimal places

# Apply bold formatting to the specified column (for PURPOSE)
    if bold_column:
        for row in ws.iter_rows(min_col=bold_column, max_col=bold_column):
            for cell in row:
                if not isinstance(cell, MergedCell):
                    continue
                cell.font = Font(name=font_name, size=font_size, bold=True)  


# Function to set column widths based on content
def auto_adjust_column_width(ws, min_row, max_row, min_col, max_col):
    for col in range(min_col, max_col + 1):
        max_length = 0
        column_letter = ws.cell(row=1, column=col).column_letter  # Get the column letter
        
        for row in range(min_row, max_row + 1):
            cell = ws.cell(row=row, column=col)
            if cell.value:
                # Calculate the maximum length of the value in the column
                max_length = max(max_length, len(str(cell.value)))

        # Set the width with some padding
        adjusted_width = max_length + 2  # Add padding for better readability
        ws.column_dimensions[column_letter].width = adjusted_width


def adjust_cell_height(ws, min_row, max_row, min_col, max_col, default_char_width = 7):

    for row_h in range(min_row, max_row + 1):
        max_lines = 1  # Reset max_lines for each row

        for col_h in range(min_col, max_col + 1):
            cell_h = ws.cell(row=row_h, column=col_h)
            if cell_h.value:
                text = str(cell_h.value)
                # Get column width, defaulting to `default_char_width` if not set
                column_letter = get_column_letter(col_h)
                col_width = ws.column_dimensions[column_letter].width or default_char_width

                # Calculate maximum line length for wrapping
                max_line_length = int(col_width / default_char_width)
                if max_line_length == 0:  # Avoid division by zero
                    max_line_length = 1

                # Count explicit lines and wrapped lines
                wrapped_lines = sum((len(line) // max_line_length) + 1 for line in text.splitlines())
                max_lines = max(max_lines, wrapped_lines)

        # Adjust row height based on the total number of lines
        ws.row_dimensions[row_h].height = max_lines * 6

# Function fo fill borders
def create_border(ws, start_col, end_col):
    cell_range_body = ws[f'B{start_col}':f'I{end_col}']
    cell_range_left = ws[f'A{start_col}':f'A{end_col}']
    cell_range_right = ws[f'J{start_col}':f'J{end_col}']

    border_body = Border(left = Side(style = 'thin'),
                    right = Side(style = 'thin'),
                    top = Side(style = 'thin'),
                    bottom = Side(style = 'thin'))
    
    border_outer_left = Border(left = Side(style = 'medium'),
                    right = Side(style = 'thin'),
                    top = Side(style = 'thin'),
                    bottom = Side(style = 'thin'))

    border_outer_right = Border(left = Side(style = 'thin'),
                    right = Side(style = 'medium'),
                    top = Side(style = 'thin'),
                    bottom = Side(style = 'thin'))
    
    for row_body in cell_range_body:
        for cell_body in row_body:
            cell_body.border = border_body
    
    for row_left in cell_range_left:
        for cell_left in row_left:
            cell_left.border = border_outer_left

    for row_right in cell_range_right:
        for cell_right in row_right:
            cell_right.border = border_outer_right

def merge_name_cell(ws, start_row, start_column, end_row, end_column, pdf_file,START_COL,END_COL):
    match = re.search(r"(\w+)-(\w+ \w+)\s+\((\w+)\s+(\d{1,2}),\s+(\d{4})\)", pdf_file)
    if match:
        last_name = match.group(1).upper()
        first_name = match.group(2).upper()
        month_str = match.group(3)
        year = match.group(5)

        # Get today's day number
        today_day = datetime.today().strftime("%d")

    ws.merge_cells(start_row=start_row, start_column=start_column, end_row=end_row, end_column=end_column)
    ws.cell(row=start_row, column=start_column).value = f"{first_name} {last_name}\n474"
    ws.cell(row=start_row, column=start_column).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws['H7'] = f"{today_day}-{month_str[:3]}-{year[-2:]}"
    ws[f"E{END_COL+1}"] = f"=SUM(E{START_COL-1}:E{END_COL-1})"
    ws[f"F{END_COL+1}"] = f"=E{END_COL+1}"

# Extract details from the `pdf_file` name
def generate_output_filename(pdf_file, start_day, end_day, start_month, end_month, end_year):
    match = re.search(r"(\w+)-(\w+ \w+)\s+\((\w+)\s+(\d{1,2}),\s+(\d{4})\)", pdf_file)
    if match:
        last_name = match.group(1)
        first_name = match.group(2)

        if start_month == end_month:
            # Combine into desired format
            output_filename = f"Reimbursement_{first_name} {last_name}_{end_month}.{start_day}-{end_day}.{end_year[-2:]}.xlsx"
        else:
            output_filename = f"Reimbursement_{first_name} {last_name}_{start_month}.{start_day}-{end_month}.{end_day}.{end_year[-2:]}.xlsx"
    
        return output_filename
    else:
        raise ValueError("Filename format does not match expected pattern.")